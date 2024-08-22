import os
import pandas as pd
from openpyxl import load_workbook
from rc2_solver_tcpc import TeamCompositionSolver, read_data
from cpsat_solver_tcpc import TeamCompositionCPSATSolver


def run_and_export(data_directory, output_file = "results.xlsx"):
    """
    Runs the solver on all .txt files in the specified directory.

    Args:
        data_directory (str): Path to the directory containing the input files.

    Returns:
        list: A list of dictionaries containing the results for each file.
    """
    # Iterate over all files in the directory
    for filename in os.listdir(data_directory):
        if filename.endswith(".txt"):
            filepath = os.path.join(data_directory, filename)
            result = run_on_file(filepath)
            print("Done results for", filename)
            export_to_excel([result], output_file)
def run_on_file(filepath):
    """
    Processes a single file and runs both RC2 and CP-SAT solvers.

    Args:
        filepath (str): Path to the input file.

    Returns:
        dict: A dictionary containing results from both solvers.
    """
    # Reading data from file
    num_students, preferences = read_data(filepath)

    # RC2 Solver with minimizing encoding
    rc2_solver_min = TeamCompositionSolver(num_students, preferences, encoding_type='min')
    rc2_assigned_min, rc2_elapsed_min = rc2_solver_min.solve()
    rc2_stats_min = rc2_solver_min.get_stats()
    print("RC2 Solver (minimizing) done")

    # CP-SAT Solver with maximizing encoding
    cpsat_solver_max = TeamCompositionCPSATSolver(num_students, preferences, encoding_type='max')
    cpsat_assigned_max, cpsat_elapsed_max = cpsat_solver_max.solve()
    cpsat_stats_max = cpsat_solver_max.get_stats()
    print("CP-SAT Solver (maximizing) done")

    # CP-SAT Solver with minimizing encoding
    cpsat_solver_min = TeamCompositionCPSATSolver(num_students, preferences, encoding_type='min')
    cpsat_assigned_min, cpsat_elapsed_min = cpsat_solver_min.solve()
    cpsat_stats_min = cpsat_solver_min.get_stats()
    print("CP-SAT Solver (minimizing) done")

    # Collecting results in a structured format
    result = {
        'num_students': num_students,
        'hard_count_rc2': rc2_stats_min['hard_clauses'],
        'hard_count_cpsat': cpsat_stats_max['hard_clauses'],
        'variables_rc2': rc2_stats_min['variables'],
        'variables_cpsat': cpsat_stats_max['variables'],
        'soft_count_min_rc2': rc2_stats_min['soft_clauses'],
        'soft_count_max_cpsat': cpsat_stats_max['soft_clauses'],
        'soft_count_min_cpsat': cpsat_stats_min['soft_clauses'],
        'time_min_rc2': rc2_elapsed_min,
        'time_max_cpsat': cpsat_elapsed_max,
        'time_min_cpsat': cpsat_elapsed_min,
    }

    return result


def export_to_excel(results, output_file):
    """
    Exports results to an Excel file. If the file already exists, it appends the new data.

    Args:
        results (list): A list of dictionaries containing solver results.
        output_file (str): Path to the output Excel file.
    """
    df = pd.DataFrame(results)

    if os.path.exists(output_file):
        # Append data to an existing file
        with pd.ExcelWriter(output_file, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
            book = load_workbook(output_file)
            sheet = book.active
            start_row = sheet.max_row  # Append data at the first available empty row
            df.to_excel(writer, index=False, startrow=start_row, header=False)
    else:
        # Create a new Excel file
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)

    print(f"Results have been appended to {output_file}")


if __name__ == "__main__":
    # Specify the directory containing input files and the output Excel file
    data_directory = 'data/temp/'
    output_file = 'results_2.xlsx'

    # Run solvers on all files in the directory and export results to Excel
    run_and_export(data_directory, output_file)
