import os
import time
import pandas as pd
from openpyxl import load_workbook
from pysat.formula import CNF, IDPool
from pysat.card import CardEnc, EncType
from pysat.solvers import Minisat22


class TeamCompositionSATSolver:
    def __init__(self, num_students, preferences):
        self.num_students = num_students
        self.preferences = preferences
        self.formula = CNF()
        self.vpool = IDPool(start_from=1)  # ID Pool for managing variables
        self.xij_vars = {}
        self.xijk_vars = {}
        self.y_vars = {}
        self.clauses_count = 0  # Biến đếm số mệnh đề
        self.solve_time = 0  # Biến lưu thời gian giải
        self.solution_found = False  # Biến lưu trạng thái của bài toán
        self.assigned_tables = []  # Biến lưu các bàn đã được sắp xếp
        self._initialize_variables()

    def _initialize_variables(self):
        """ Initialize Boolean variables for the formula. """
        for i in range(1, self.num_students + 1):
            for j in range(i + 1, self.num_students + 1):
                self.xij_vars[(i, j)] = self.vpool.id()
            for j in range(i + 1, self.num_students + 1):
                for k in range(j + 1, self.num_students + 1):
                    self.xijk_vars[(i, j, k)] = self.vpool.id()
            self.y_vars[i] = self.vpool.id()

    def add_hard_clauses(self):
        """ Add hard constraints to the formula. """
        self._add_single_assignment_clauses()
        self._add_valid_table_clauses()
        self._add_cardinality_constraint()

    def _add_single_assignment_clauses(self):
        """ Add constraints ensuring each student is assigned to exactly one table. """
        for i in range(1, self.num_students + 1):
            clause = self._get_single_assignment_clause(i)
            card_enc_atmost = CardEnc.atmost(lits=clause, bound=1, vpool=self.vpool, encoding=EncType.seqcounter)
            for c in card_enc_atmost.clauses:
                self.formula.append(c)
            self.formula.append(clause)
            self.clauses_count += 1  # Tăng số lượng mệnh đề

    def _get_single_assignment_clause(self, i):
        """ Generate clause for single assignment of student i. """
        clause = []
        if i == 1:
            clause = [self.xij_vars[(i, j)] for j in range(2, self.num_students + 1)]
            clause += [self.xijk_vars[(i, j, k)] for j in range(2, self.num_students) for k in range(j + 1, self.num_students + 1)]
        elif 2 <= i <= self.num_students - 1:
            clause = [self.xij_vars[(j, i)] for j in range(1, i)] + [self.xij_vars[(i, j)] for j in range(i + 1, self.num_students + 1)]
            clause += [self.xijk_vars[(j, k, i)] for k in range(2, i) for j in range(1, k)]
            clause += [self.xijk_vars[(j, i, k)] for j in range(1, i) for k in range(i + 1, self.num_students + 1)]
            clause += [self.xijk_vars[(i, j, k)] for j in range(i + 1, self.num_students) for k in range(j + 1, self.num_students + 1)]
        else:  # i == num_students
            clause = [self.xij_vars[(j, i)] for j in range(1, self.num_students)]
            clause += [self.xijk_vars[(j, k, i)] for j in range(1, self.num_students - 1) for k in range(j + 1, self.num_students)]
        return clause

    def _add_valid_table_clauses(self):
        """ Add constraints ensuring valid table assignments. """
        for (i, j) in self.xij_vars:
            self.formula.append([-self.xij_vars[(i, j)], self.y_vars[i]])
            self.formula.append([-self.xij_vars[(i, j)], self.y_vars[j]])
            self.clauses_count += 2  # Tăng số lượng mệnh đề

        for (i, j, k) in self.xijk_vars:
            self.formula.append([-self.xijk_vars[(i, j, k)], -self.y_vars[i]])
            self.formula.append([-self.xijk_vars[(i, j, k)], -self.y_vars[j]])
            self.formula.append([-self.xijk_vars[(i, j, k)], -self.y_vars[k]])
            self.clauses_count += 3  # Tăng số lượng mệnh đề

    def _add_cardinality_constraint(self):
        """ Add cardinality constraints for the number of tables. """
        num_tables_2 = int(self.num_students * 4 / 7)
        card_constraint = CardEnc.equals(lits=list(self.y_vars.values()), bound=num_tables_2, vpool=self.vpool, encoding=EncType.seqcounter)
        for clause in card_constraint.clauses:
            self.formula.append(clause)
            self.clauses_count += 1  # Tăng số lượng mệnh đề

    def calculate_weights(self):
        """ Calculate weights based on the chosen encoding type. """
        out_degrees = {i: 0 for i in range(1, self.num_students + 1)}
        for i, friends in self.preferences.items():
            out_degrees[i] = len(friends)

        wij = {}
        wijk = {}

        for i in range(1, self.num_students + 1):
            for j in range(i + 1, self.num_students + 1):
                subgraph_out_degrees = {
                    v: sum(1 for neighbor in self.preferences.get(v, []) if neighbor in {i, j})
                    for v in [i, j]
                }
                wi = subgraph_out_degrees[i]
                wj = subgraph_out_degrees[j]
                wij[(i, j)] = 2 * wi * wj
            for j in range(i + 1, self.num_students + 1):
                for k in range(j + 1, self.num_students + 1):
                    subgraph_out_degrees = {
                        v: sum(1 for neighbor in self.preferences.get(v, []) if neighbor in {i, j, k})
                        for v in [i, j, k]
                    }
                    wi = subgraph_out_degrees[i]
                    wj = subgraph_out_degrees[j]
                    wk = subgraph_out_degrees[k]
                    wijk[(i, j, k)] = 3 * wi * wj * wk / 8
        return wij, wijk

    def add_constraint_through_preferences(self, wij, wijk):
        for (i, j), weight in wij.items():
            if weight != 2:
                self.formula.append([-self.xij_vars[(i, j)]])
                self.clauses_count += 1  # Tăng số lượng mệnh đề

        for (i, j, k), weight in wijk.items():
            if weight != 3:
                self.formula.append([-self.xijk_vars[(i, j, k)]])
                self.clauses_count += 1  # Tăng số lượng mệnh đề

    def solve(self):
        """ Giải bài toán và trả về kết quả """
        self.add_hard_clauses()
        wij, wijk = self.calculate_weights()
        self.add_constraint_through_preferences(wij, wijk)
        solver = Minisat22(bootstrap_with=self.formula.clauses)
        start_time = time.time()
        self.solution_found = solver.solve()
        self.solve_time = time.time() - start_time

        # Trích xuất mô hình (model) nếu bài toán SAT thỏa mãn
        model = solver.get_model() if self.solution_found else None
        if model:
            self.assigned_tables = self.extract_solution(model)

    def extract_solution(self, model):
        """ Extract the assigned tables from the solution """
        assigned_tables = {}

        # Tạo từ điển để tra cứu nhanh biến
        var_to_students_pair = {v: (i, j) for (i, j), v in self.xij_vars.items()}
        var_to_students_triple = {v: (i, j, k) for (i, j, k), v in self.xijk_vars.items()}

        for var in model:
            if var > 0:  # Chỉ xem xét những biến dương (được chọn trong solution)
                if var in var_to_students_pair:
                    i, j = var_to_students_pair[var]
                    assigned_tables.setdefault(var, []).extend([i, j])
                elif var in var_to_students_triple:
                    i, j, k = var_to_students_triple[var]
                    assigned_tables.setdefault(var, []).extend([i, j, k])

        # Loại bỏ trùng lặp trong các bảng
        final_assigned_tables = [list(set(students)) for students in assigned_tables.values()]

        return final_assigned_tables

    def get_stats(self):
        """ Trả về số lượng biến, số lượng mệnh đề, và trạng thái bài toán """
        num_variables = len(self.xij_vars) + len(self.xijk_vars) + len(self.y_vars)
        return {
            'variables': num_variables,  # Tổng số biến
            'clauses': self.clauses_count,  # Số lượng mệnh đề đã thêm
            'solve_time': self.solve_time,  # Thời gian giải bài toán
            'solution_found': self.solution_found  # Bài toán có giải được không?
        }

    def print_assigned_tables(self):
        """ In ra danh sách các bàn đã được sắp xếp """
        print("Assigned tables:")
        for table in self.assigned_tables:
            print(table)


def read_data(filename):
    """ Đọc dữ liệu từ file .txt """
    with open(filename, 'r') as file:
        lines = file.readlines()

    num_students = int(lines[0].strip())
    preferences = {}
    for line in lines[1:]:
        parts = list(map(int, line.strip().split()))
        preferences[parts[0]] = parts[1:]

    return num_students, preferences


def export_to_excel(results, output_file):
    """
    Exports results to an Excel file. If the file already exists, it appends the new data.

    Args:
        results (list): A list of dictionaries containing solver results.
        output_file (str): Path to the output Excel file.
    """
    df = pd.DataFrame(results)

    if os.path.exists(output_file):
        # Append data to an existing file
        with pd.ExcelWriter(output_file, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
            book = load_workbook(output_file)
            sheet = book.active
            start_row = sheet.max_row  # Append data at the first available empty row
            df.to_excel(writer, index=False, startrow=start_row, header=False)
    else:
        # Create a new Excel file
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)

    print(f"Results have been appended to {output_file}")


def run_and_export(data_directory, output_file="results.xlsx", num_runs=1):
    """
    Runs the solver on all .txt files in the specified directory multiple times and averages the results.

    Args:
        data_directory (str): Path to the directory containing the input files.
        output_file (str): Path to the output Excel file.
        num_runs (int): Number of times to run the solver to average the time and total weight.
    """
    for filename in os.listdir(data_directory):
        if filename.endswith(".txt"):
            filepath = os.path.join(data_directory, filename)
            print(f"Running on {filename}")
            result = run_on_file(filepath, num_runs)
            print(f"Done results for {filename}")
            export_to_excel([result], output_file)


def run_on_file(filepath, num_runs=1):
    """
    Processes a single file and runs TeamCompositionSolver multiple times to average the time.

    Args:
        filepath (str): Path to the input file.
        num_runs (int): Number of times to run the solver to average the time.

    Returns:
        dict: A dictionary containing averaged results from the solver, including the filename.
    """
    # Reading data from file
    num_students, preferences = read_data(filepath)

    # Initialize variables to store cumulative results
    total_time = 0
    solver_stats = None

    # Run the solver multiple times and accumulate the results
    for _ in range(num_runs):
        solver = TeamCompositionSATSolver(num_students, preferences)
        solver.solve()
        stats = solver.get_stats()

        total_time += stats['solve_time']
        solver_stats = stats  # Keep updating to store the final stats (they should be the same across runs)

    # Calculate average time
    avg_time = total_time / num_runs
    filename = os.path.basename(filepath)  # Extract filename for reporting

    # Create the result dictionary
    result = {
        'filename': filename,
        'num_students': num_students,
        'variables': solver_stats['variables'],
        'clauses': solver_stats['clauses'],
        'avg_time': avg_time,
        'solution_found': 'SAT' if solver_stats['solution_found'] else 'UNSAT'
    }

    return result


if __name__ == "__main__":
    # Specify the directory containing input files and the output Excel file
    data_directory = 'data/max/'
    output_file = 'results_solver_sss.xlsx'

    # Run solvers on all files in the directory and export results to Excel
    run_and_export(data_directory, output_file)
