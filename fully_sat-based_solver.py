import os
import time
import pandas as pd
from openpyxl.reader.excel import load_workbook
from pysat.formula import CNF, IDPool
from pysat.card import CardEnc, EncType
from pysat.solvers import Minisat22


class TeamCompositionSolver:
    def __init__(self, num_students, preferences):
        self.num_students = num_students
        self.preferences = preferences
        self.formula = CNF()
        self.vpool = IDPool(start_from=1)  # ID Pool for managing variables
        self.xij_vars = {}
        self.xijk_vars = {}
        self.y_vars = {}
        self._initialize_variables()
        self.clauses_count = 0  # Biến đếm số mệnh đề

    def _initialize_variables(self):
        """ Initialize Boolean variables for the formula. """
        for i in range(1, self.num_students + 1):
            for j in range(i + 1, self.num_students + 1):
                self.xij_vars[(i, j)] = self.vpool.id()
            for j in range(i + 1, self.num_students + 1):
                for k in range(j + 1, self.num_students + 1):
                    self.xijk_vars[(i, j, k)] = self.vpool.id()
            self.y_vars[i] = self.vpool.id()

    def add_hard_clauses(self):
        """ Add hard constraints to the formula. """
        self._add_single_assignment_clauses()
        self._add_valid_table_clauses()
        self._add_cardinality_constraint()

    def _add_single_assignment_clauses(self):
        """ Add constraints ensuring each student is assigned to exactly one table. """
        for i in range(1, self.num_students + 1):
            clause = self._get_single_assignment_clause(i)
            card_enc_atmost = CardEnc.atmost(lits=clause, bound=1, vpool=self.vpool, encoding=EncType.seqcounter)
            for c in card_enc_atmost.clauses:
                self.formula.append(c)
            self.formula.append(clause)
            self.clauses_count += 1  # Tăng số lượng mệnh đề

    def _get_single_assignment_clause(self, i):
        """ Generate clause for single assignment of student i. """
        clause = []
        if i == 1:
            clause = [self.xij_vars[(i, j)] for j in range(2, self.num_students + 1)]
            clause += [self.xijk_vars[(i, j, k)] for j in range(2, self.num_students) for k in range(j + 1, self.num_students + 1)]
        elif 2 <= i <= self.num_students - 1:
            clause = [self.xij_vars[(j, i)] for j in range(1, i)] + [self.xij_vars[(i, j)] for j in range(i + 1, self.num_students + 1)]
            clause += [self.xijk_vars[(j, k, i)] for k in range(2, i) for j in range(1, k)]
            clause += [self.xijk_vars[(j, i, k)] for j in range(1, i) for k in range(i + 1, self.num_students + 1)]
            clause += [self.xijk_vars[(i, j, k)] for j in range(i + 1, self.num_students) for k in range(j + 1, self.num_students + 1)]
        else:  # i == num_students
            clause = [self.xij_vars[(j, i)] for j in range(1, self.num_students)]
            clause += [self.xijk_vars[(j, k, i)] for j in range(1, self.num_students - 1) for k in range(j + 1, self.num_students)]
        return clause

    def _add_valid_table_clauses(self):
        """ Add constraints ensuring valid table assignments. """
        for (i, j) in self.xij_vars:
            self.formula.append([-self.xij_vars[(i, j)], self.y_vars[i]])
            self.formula.append([-self.xij_vars[(i, j)], self.y_vars[j]])
            self.clauses_count += 2  # Tăng số lượng mệnh đề

        for (i, j, k) in self.xijk_vars:
            self.formula.append([-self.xijk_vars[(i, j, k)], -self.y_vars[i]])
            self.formula.append([-self.xijk_vars[(i, j, k)], -self.y_vars[j]])
            self.formula.append([-self.xijk_vars[(i, j, k)], -self.y_vars[k]])
            self.clauses_count += 3  # Tăng số lượng mệnh đề

    def _add_cardinality_constraint(self):
        """ Add cardinality constraints for the number of tables. """
        num_tables_2 = int(self.num_students * 4 / 7)
        card_constraint = CardEnc.equals(lits=list(self.y_vars.values()), bound=num_tables_2, vpool=self.vpool, encoding=EncType.seqcounter)
        for clause in card_constraint.clauses:
            self.formula.append(clause)
            self.clauses_count += 1  # Tăng số lượng mệnh đề

    def calculate_weights(self):
        """ Calculate weights based on the chosen encoding type. """
        out_degrees = {i: 0 for i in range(1, self.num_students + 1)}
        for i, friends in self.preferences.items():
            out_degrees[i] = len(friends)

        wij = {}
        wijk = {}

        for i in range(1, self.num_students + 1):
            for j in range(i + 1, self.num_students + 1):
                subgraph_out_degrees = {
                    v: sum(1 for neighbor in self.preferences.get(v, []) if neighbor in {i, j})
                    for v in [i, j]
                }
                wi = subgraph_out_degrees[i]
                wj = subgraph_out_degrees[j]
                wij[(i, j)] = 2 * wi * wj
            for j in range(i + 1, self.num_students + 1):
                for k in range(j + 1, self.num_students + 1):
                    subgraph_out_degrees = {
                        v: sum(1 for neighbor in self.preferences.get(v, []) if neighbor in {i, j, k})
                        for v in [i, j, k]
                    }
                    wi = subgraph_out_degrees[i]
                    wj = subgraph_out_degrees[j]
                    wk = subgraph_out_degrees[k]
                    wijk[(i, j, k)] = 3 * wi * wj * wk / 8
        return wij, wijk

    def add_constraint_through_preferences(self, wij, wijk):
        for (i, j), weight in wij.items():
            if weight != 2:
                self.formula.append([-self.xij_vars[(i, j)]])
                self.clauses_count += 1  # Tăng số lượng mệnh đề

        for (i, j, k), weight in wijk.items():
            if weight != 3:
                self.formula.append([-self.xijk_vars[(i, j, k)]])
                self.clauses_count += 1  # Tăng số lượng mệnh đề
    def solve(self):
        """ Giải bài toán và trả về kết quả """
        self.add_hard_clauses()
        wij, wijk = self.calculate_weights()
        self.add_constraint_through_preferences(wij, wijk)

        solver = Minisat22(bootstrap_with=self.formula.clauses)
        start_time = time.time()
        solution_found = solver.solve()
        elapsed_time = time.time() - start_time

        # Trích xuất mô hình (model) nếu bài toán SAT thỏa mãn
        model = solver.get_model() if solution_found else None
        return model, elapsed_time, solution_found

    def extract_solution_and_calculate_weights(self, model):
        """ Extract the assigned tables from the solution and calculate the total satisfied weights. """
        assigned_tables = {}
        total_satisfied_weight = 0

        if not model:
            return [], 0

        # Tạo từ điển để tra cứu nhanh biến
        var_to_students_pair = {v: (i, j) for (i, j), v in self.xij_vars.items()}
        var_to_students_triple = {v: (i, j, k) for (i, j, k), v in self.xijk_vars.items()}

        for var in model:
            if var > 0:  # Chỉ xem xét những biến dương (được chọn trong solution)
                if var in var_to_students_pair:
                    i, j = var_to_students_pair[var]
                    assigned_tables.setdefault(var, []).extend([i, j])
                    total_satisfied_weight += 2  # Bạn có thể cập nhật cách tính trọng số nếu cần
                elif var in var_to_students_triple:
                    i, j, k = var_to_students_triple[var]
                    assigned_tables.setdefault(var, []).extend([i, j, k])
                    total_satisfied_weight += 3  # Bạn có thể cập nhật cách tính trọng số nếu cần

        # Loại bỏ trùng lặp trong các bảng
        final_assigned_tables = [list(set(students)) for students in assigned_tables.values()]

        return final_assigned_tables, total_satisfied_weight

    def get_stats(self):
        """ Trả về số lượng biến và số lượng mệnh đề """
        num_variables = len(self.xij_vars) + len(self.xijk_vars) + len(self.y_vars)
        return {
            'variables': num_variables,  # Tổng số biến
            'clauses': self.clauses_count,  # Số lượng mệnh đề đã thêm
        }


def read_data(filename):
    """ Đọc dữ liệu từ file .txt """
    with open(filename, 'r') as file:
        lines = file.readlines()

    num_students = int(lines[0].strip())
    preferences = {}
    for line in lines[1:]:
        parts = list(map(int, line.strip().split()))
        preferences[parts[0]] = parts[1:]

    return num_students, preferences


def export_to_excel(results, output_file):
    """
    Exports results to an Excel file. If the file already exists, it appends the new data.

    Args:
        results (list): A list of dictionaries containing solver results.
        output_file (str): Path to the output Excel file.
    """
    df = pd.DataFrame(results)

    if os.path.exists(output_file):
        # Append data to an existing file
        with pd.ExcelWriter(output_file, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
            book = load_workbook(output_file)
            sheet = book.active
            start_row = sheet.max_row  # Append data at the first available empty row
            df.to_excel(writer, index=False, startrow=start_row, header=False)
    else:
        # Create a new Excel file
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)

    print(f"Results have been appended to {output_file}")


if __name__ == "__main__":
    # input_data = 'data/max/students_preferences_126.txt'
    # num_students, preferences = read_data(input_data)
    #
    # solver = TeamCompositionSolver(num_students, preferences)
    # solution, elapsed_time, solution_found = solver.solve()
    # # final_assigned_tables, total_satisfied_weight = solver.extract_solution_and_calculate_weights(solution)
    # stats = solver.get_stats()
    #
    # # Print results
    # print(f"Students: {num_students}")
    # print(f"Variables: {stats['variables']}")
    # print(f"Clauses: {stats['clauses']}")
    # print(f"Time for solving: {elapsed_time:.2f} seconds")
    # if solution:
    #     print("Solution found!")
    #     # print("Assigned tables:")
    #     # for table in solution:
    #     #     print(table)
    #     # print(f"Total satisfied weight: {total_satisfied_weight}")
    # else:
    #     print("No solution found!")

    # Đọc dữ liệu từ thư mục chứa các file input
    data_directory = 'data/'
    output_file = 'sat_results_2.xlsx'
    results = []

    # Lặp qua tất cả các file .txt trong thư mục
    for filename in os.listdir(data_directory):
        if filename.endswith(".txt"):
            filepath = os.path.join(data_directory, filename)
            num_students, preferences = read_data(filepath)

            solver = TeamCompositionSolver(num_students, preferences)
            model, elapsed_time, solution_found = solver.solve()
            stats = solver.get_stats()

            # Lưu kết quả
            result = {
                'num_students': num_students,
                'variables': stats['variables'],
                'clauses': stats['clauses'],
                'time': elapsed_time,
                'result': 'SAT' if solution_found else 'UNSAT'
            }
            export_to_excel([result], output_file)
            print(f"Đã hoàn thành: {filename}")

