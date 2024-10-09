import os
import pandas as pd
from openpyxl import load_workbook
from rc2_solver_timeout import TeamCompositionSolver, read_data



def run_and_export(data_directory, output_file="results.xlsx", num_runs=2):
    """
    Runs the solver on all .txt files in the specified directory multiple times and averages the results.

    Args:
        data_directory (str): Path to the directory containing the input files.

    Returns:
        list: A list of dictionaries containing the results for each file.
    """
    for filename in os.listdir(data_directory):
        if filename.endswith(".txt"):
            filepath = os.path.join(data_directory, filename)
            print("Running on", filename)
            result = run_on_file(filepath, num_runs=num_runs)
            print("Done results for", filename)
            export_to_excel([result], output_file)


def run_on_file(filepath, num_runs=1):
    """
    Processes a single file and runs the RC2 solver multiple times with maximizing encoding to average the time and total weight.

    Args:
        filepath (str): Path to the input file.
        num_runs (int): Number of times to run the solver to average the time and total weight.

    Returns:
        dict: A dictionary containing averaged results from the RC2 solver, including the filename.
    """
    # Reading data from file
    num_students, preferences = read_data(filepath)

    # RC2 Solver with maximizing encoding
    total_time_rc2_max, total_weight_rc2_max = 0, 0
    rc2_stats_max = None  # Khởi tạo rc2_stats_max để lưu kết quả cuối cùng

    for _ in range(num_runs):
        rc2_solver_max = TeamCompositionSolver(num_students, preferences, encoding_type='max')
        rc2_solver_max.solve()
        rc2_stats_max = rc2_solver_max.get_stats()  # Lấy kết quả sau mỗi lần chạy
        total_time_rc2_max += rc2_stats_max['solve_time']
        total_weight_rc2_max += rc2_stats_max['total_weight']

    # Tính thời gian và trọng số trung bình sau num_runs lần chạy
    avg_time_rc2_max = total_time_rc2_max / num_runs
    avg_weight_rc2_max = total_weight_rc2_max / num_runs
    print("RC2 Solver (maximizing) done")

    # Extracting the filename from filepath
    filename = os.path.basename(filepath)

    # Collecting results in a structured format to return
    result = {
        'filename': filename,  # Thêm tên file vào kết quả
        'num_students': num_students,
        'hard_count_rc2': rc2_stats_max['hard_clauses'],
        'variables_rc2': rc2_stats_max['variables'],
        'soft_count_max_rc2': rc2_stats_max['soft_clauses'],
        'time_max_rc2': avg_time_rc2_max,
        'total_weight_max_rc2': avg_weight_rc2_max
    }

    return result




def export_to_excel(results, output_file):
    """
    Exports results to an Excel file. If the file already exists, it appends the new data.

    Args:
        results (list): A list of dictionaries containing solver results.
        output_file (str): Path to the output Excel file.
    """
    df = pd.DataFrame(results)

    if os.path.exists(output_file):
        # Append data to an existing file
        with pd.ExcelWriter(output_file, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
            book = load_workbook(output_file)
            sheet = book.active
            start_row = sheet.max_row  # Append data at the first available empty row
            df.to_excel(writer, index=False, startrow=start_row, header=False)
    else:
        # Create a new Excel file
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)

    print(f"Results have been appended to {output_file}")


if __name__ == "__main__":
    # # Specify the directory containing input files and the output Excel file
    # data_directory = 'data/fully/'
    # output_file = 'results/fully.xlsx'
    #
    # # Run solvers on all files in the directory and export results to Excel
    # run_and_export(data_directory, output_file)
    #
    # data_directory = 'data/max/max1'
    # output_file = 'results/max.xlsx'
    # run_and_export(data_directory, output_file)

    # data_directory = 'data/max/max2'
    # output_file = 'results/max.xlsx'
    # run_and_export(data_directory, output_file, 1)

    data_directory = 'data/temp/'
    output_file = 'results/temp.xlsx'
    run_and_export(data_directory, output_file)
