import os
import pandas as pd
from openpyxl import load_workbook
from rc2_solver_tcpc import TeamCompositionSolver, read_data
from cpsat_solver import TeamCompositionCPSATSolver
from sat_solver import TeamCompositionSATSolver


def run_and_export(data_directory, output_file="results.xlsx", num_runs=2):
    """
    Runs the solver on all .txt files in the specified directory multiple times and averages the results.

    Args:
        data_directory (str): Path to the directory containing the input files.

    Returns:
        list: A list of dictionaries containing the results for each file.
    """
    for filename in os.listdir(data_directory):
        if filename.endswith(".txt"):
            filepath = os.path.join(data_directory, filename)
            print("Running on", filename)
            result = run_on_file(filepath, num_runs=num_runs)
            print("Done results for", filename)
            export_to_excel([result], output_file)


def run_on_file(filepath, num_runs=2):
    """
    Processes a single file and runs both RC2 and CP-SAT solvers multiple times to average the time and total weight.

    Args:
        filepath (str): Path to the input file.
        num_runs (int): Number of times to run the solvers to average the time and total weight.

    Returns:
        dict: A dictionary containing averaged results from both solvers, including the filename.
    """
    # Reading data from file
    num_students, preferences = read_data(filepath)

    ### SAT Solver ###
    total_time_sat, solution_found = 0, False
    sat_stats = None

    for _ in range(num_runs):
        sat_solver = TeamCompositionSATSolver(num_students, preferences)
        sat_solver.solve()
        sat_stats = sat_solver.get_stats()
        total_time_sat += sat_stats['solve_time']
        solution_found = sat_stats['solution_found']  # Keep track of solution existence
    avg_time_sat = total_time_sat / num_runs
    print("SAT Solver done")

    # RC2 Solver with minimizing encoding
    total_time_rc2_min, total_weight_rc2_min = 0, 0
    rc2_stats_min = None  # Khởi tạo rc2_stats_min để lưu kết quả cuối cùng

    for _ in range(num_runs):
        rc2_solver_min = TeamCompositionSolver(num_students, preferences, encoding_type='min')
        rc2_solver_min.solve()
        rc2_stats_min = rc2_solver_min.get_stats()  # Lấy kết quả sau mỗi lần chạy
        total_time_rc2_min += rc2_stats_min['solve_time']
        total_weight_rc2_min += rc2_stats_min['total_weight']
    avg_time_rc2_min = total_time_rc2_min / num_runs
    avg_weight_rc2_min = total_weight_rc2_min / num_runs
    print("RC2 Solver (minimizing) done")

    # CP-SAT Solver with maximizing encoding
    total_time_cpsat_max, total_weight_cpsat_max = 0, 0
    cpsat_stats_max = None  # Khởi tạo cpsat_stats_max để lưu kết quả cuối cùng

    for _ in range(num_runs):
        cpsat_solver_max = TeamCompositionCPSATSolver(num_students, preferences, encoding_type='max')
        cpsat_solver_max.solve()
        cpsat_stats_max = cpsat_solver_max.get_stats()  # Lấy kết quả sau mỗi lần chạy
        total_time_cpsat_max += cpsat_stats_max['solve_time']
        total_weight_cpsat_max += cpsat_stats_max['total_weight']
    avg_time_cpsat_max = total_time_cpsat_max / num_runs
    avg_weight_cpsat_max = total_weight_cpsat_max / num_runs
    print("CP-SAT Solver (maximizing) done")

    # CP-SAT Solver with minimizing encoding
    total_time_cpsat_min, total_weight_cpsat_min = 0, 0
    cpsat_stats_min = None  # Khởi tạo cpsat_stats_min để lưu kết quả cuối cùng

    for _ in range(num_runs):
        cpsat_solver_min = TeamCompositionCPSATSolver(num_students, preferences, encoding_type='min')
        cpsat_solver_min.solve()
        cpsat_stats_min = cpsat_solver_min.get_stats()  # Lấy kết quả sau mỗi lần chạy
        total_time_cpsat_min += cpsat_stats_min['solve_time']
        total_weight_cpsat_min += cpsat_stats_min['total_weight']
    avg_time_cpsat_min = total_time_cpsat_min / num_runs
    avg_weight_cpsat_min = total_weight_cpsat_min / num_runs
    print("CP-SAT Solver (minimizing) done")

    # Extracting the filename from filepath
    filename = os.path.basename(filepath)

    # Collecting results in a structured format to return
    result = {
        'filename': filename,  # Thêm tên file vào kết quả
        'num_students': num_students,
        'sat_clauses': sat_stats['clauses'],  # SAT solver clauses
        'time_sat': avg_time_sat,  # SAT solver average time
        'solution_found': 'SAT' if sat_stats['solution_found'] else 'UNSAT' , # Whether the SAT solution was found
        'hard_count_rc2': rc2_stats_min['hard_clauses'],
        'hard_count_cpsat': cpsat_stats_max['hard_clauses'],
        'variables_rc2': rc2_stats_min['variables'],
        'variables_cpsat': cpsat_stats_max['variables'],
        'soft_count_min_rc2': rc2_stats_min['soft_clauses'],
        'soft_count_max_cpsat': cpsat_stats_max['soft_clauses'],
        'soft_count_min_cpsat': cpsat_stats_min['soft_clauses'],
        'time_min_rc2': avg_time_rc2_min,
        'time_max_cpsat': avg_time_cpsat_max,
        'time_min_cpsat': avg_time_cpsat_min,
        'total_weight_min_rc2': avg_weight_rc2_min,
        'total_weight_max_cpsat': avg_weight_cpsat_max,
        'total_weight_min_cpsat': avg_weight_cpsat_min,
    }
    print(result)
    return result



def export_to_excel(results, output_file):
    """
    Exports results to an Excel file. If the file already exists, it appends the new data.

    Args:
        results (list): A list of dictionaries containing solver results.
        output_file (str): Path to the output Excel file.
    """
    df = pd.DataFrame(results)

    if os.path.exists(output_file):
        # Append data to an existing file
        with pd.ExcelWriter(output_file, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
            book = load_workbook(output_file)
            sheet = book.active
            start_row = sheet.max_row  # Append data at the first available empty row
            df.to_excel(writer, index=False, startrow=start_row, header=False)
    else:
        # Create a new Excel file
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)

    print(f"Results have been appended to {output_file}")


if __name__ == "__main__":
    # # Specify the directory containing input files and the output Excel file
    # data_directory = 'data/fully/'
    # output_file = 'results/fully.xlsx'
    #
    # # Run solvers on all files in the directory and export results to Excel
    # run_and_export(data_directory, output_file)
    #
    # data_directory = 'data/max/max1'
    # output_file = 'results/max.xlsx'
    # run_and_export(data_directory, output_file)

    # data_directory = 'data/max/max2'
    # output_file = 'results/max.xlsx'
    # run_and_export(data_directory, output_file, 1)

    data_directory = 'data/temp/'
    output_file = 'results/temp.xlsx'
    run_and_export(data_directory, output_file, 1)
